from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.db.models import Q
from .models import Producto
from .serializers import ProductoSerializer
from rest_framework.parsers import MultiPartParser
from decimal import Decimal, InvalidOperation
from django.db import transaction
import openpyxl

# Configuración de paginación
class ProductoPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def producto_por_codigo(request):
    codigo_barra = request.GET.get('codigo_barra', '')
    if not codigo_barra:
        return Response({'error': 'Se requiere codigo_barra'}, status=400)
    
    try:
        producto = Producto.objects.get(codigo_barra=codigo_barra)
        serializer = ProductoSerializer(producto)
        return Response(serializer.data)
    except Producto.DoesNotExist:
        return Response({'error': 'Producto no encontrado'}, status=404)

# ============================================
# 1. LISTAR PRODUCTOS
# ============================================
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def producto_list(request):
    productos = Producto.objects.all().order_by('-id')
    
    # Filtro por búsqueda
    search = request.GET.get('search', '')
    if search:
        productos = productos.filter(
            Q(nombre_producto__icontains=search) |
            Q(codigo_barra__icontains=search) | 
            Q(codigo_interno__icontains=search)
        )
    
    # Filtro por estado
    estado = request.GET.get('estado')
    if estado is not None:
        if estado.lower() == 'true':
            productos = productos.filter(estado=True)
        elif estado.lower() == 'false':
            productos = productos.filter(estado=False)
    
    # Filtros de precio de compra
    costo_min = request.GET.get('costo_min')
    costo_max = request.GET.get('costo_max')
    if costo_min is not None and costo_min != '':
        try:
            productos = productos.filter(costo_producto__gte=float(costo_min))
        except ValueError:
            pass
    if costo_max is not None and costo_max != '':
        try:
            productos = productos.filter(costo_producto__lte=float(costo_max))
        except ValueError:
            pass
    
    # Filtros de precio de venta
    precio_min = request.GET.get('precio_min')
    precio_max = request.GET.get('precio_max')
    if precio_min is not None and precio_min != '':
        try:
            productos = productos.filter(precio_unitario__gte=float(precio_min))
        except ValueError:
            pass
    if precio_max is not None and precio_max != '':
        try:
            productos = productos.filter(precio_unitario__lte=float(precio_max))
        except ValueError:
            pass

    # Paginación
    paginator = ProductoPagination()
    paginated_productos = paginator.paginate_queryset(productos, request)
    serializer = ProductoSerializer(paginated_productos, many=True, context={'request': request})
    
    return paginator.get_paginated_response(serializer.data)

# ============================================
# 2. OBTENER PRODUCTO POR ID
# ============================================
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def producto_detail(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    serializer = ProductoSerializer(producto, context={'request': request})
    return Response(serializer.data)

# ============================================
# 3. CREAR PRODUCTO
# ============================================
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def producto_create(request):

    serializer = ProductoSerializer(data=request.data, context={'request': request})
    
    if serializer.is_valid():
        producto = serializer.save()
        return Response(
            {
                'message': 'Producto creado exitosamente',
                'producto': ProductoSerializer(producto).data
            },
            status=status.HTTP_201_CREATED
        )
    
    return Response(
        {
            'error': 'Error al crear el producto',
            'detalles': serializer.errors
        },
        status=status.HTTP_400_BAD_REQUEST
    )

# ============================================
# 4. ACTUALIZAR PRODUCTO (TODOS LOS CAMPOS)
# ============================================
@api_view(['PUT', 'PATCH'])
@permission_classes([IsAuthenticated])
def producto_update(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    partial = request.method == 'PATCH'
    
    serializer = ProductoSerializer(
        producto, 
        data=request.data, 
        partial=partial, 
        context={'request': request}
    )
    
    if serializer.is_valid():
        producto_actualizado = serializer.save()
        return Response(
            {
                'message': 'Producto actualizado exitosamente',
                'producto': ProductoSerializer(producto_actualizado).data
            }
        )
    return Response(
        {
            'error': 'Error al actualizar el producto',
            'detalles': serializer.errors
        },
        status=status.HTTP_400_BAD_REQUEST
    )

# ============================================
# 5. ELIMINAR PRODUCTO
# ============================================
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def producto_delete(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    nombre_producto = producto.nombre_producto
    producto.delete()
    
    return Response(
        {
            'message': f'Producto "{nombre_producto}" eliminado exitosamente'
        },
        status=status.HTTP_200_OK
    )

# ============================================
# 6. IMPORTAR PRODUCTOS DESDE EXCEL
# ============================================
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def importar_productos_excel(request):

    """
    Reglas:
    - Si el código de barra NO existe: CREA el producto con todos los datos
    - Si el código de barra YA existe: SOLO actualiza el NOMBRE (NO modifica cantidad, precios ni estado)
    """
    from decimal import Decimal
    
    # Verificar si el archivo está en request.FILES
    archivo = request.FILES.get('archivo')
    
    if not archivo:
        return Response(
            {'error': 'No se proporcionó ningún archivo. El campo debe llamarse "archivo"'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Validar extensión
    if not archivo.name.endswith(('.xlsx', '.xls')):
        return Response(
            {'error': 'Formato de archivo no válido. Use .xlsx o .xls'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # Leer el archivo Excel con openpyxl
        workbook = openpyxl.load_workbook(archivo, data_only=True)
        sheet = workbook.active
        
        # Verificar que la hoja tenga datos
        if sheet.max_row < 2:
            return Response(
                {'error': 'El archivo no contiene datos (mínimo 2 filas: cabecera + datos)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        productos_creados = []
        productos_actualizados = []
        errores = []
        
        with transaction.atomic():
            # Empezar desde la fila 2 (asumiendo que fila 1 es cabecera)
            for row_idx in range(2, sheet.max_row + 1):
                try:
                    # Extraer datos de las columnas
                    nombre = sheet.cell(row=row_idx, column=1).value
                    cantidad = sheet.cell(row=row_idx, column=2).value
                    costo = sheet.cell(row=row_idx, column=3).value
                    precio = sheet.cell(row=row_idx, column=4).value
                    codigo_barra = sheet.cell(row=row_idx, column=5).value
                    
                    # Limpiar y validar datos
                    if nombre is None or str(nombre).strip() == '':
                        errores.append(f"Fila {row_idx}: El nombre del producto es requerido")
                        continue
                    
                    nombre = str(nombre).strip()
                    
                    # Manejar cantidad
                    if cantidad is None:
                        cantidad = 0
                    elif isinstance(cantidad, str):
                        cantidad_str = cantidad.replace(',', '.').strip()
                        cantidad = int(float(cantidad_str)) if cantidad_str else 0
                    else:
                        cantidad = int(cantidad)
                    
                    # Función para convertir precios
                    def convertir_precio(valor):
                        if valor is None:
                            return 0.0
                        if isinstance(valor, (int, float)):
                            return float(valor)
                        if isinstance(valor, str):
                            valor = valor.replace(',', '.').strip()
                            return float(valor) if valor else 0.0
                        return 0.0
                    
                    costo = convertir_precio(costo)
                    precio = convertir_precio(precio)
                    
                    # Manejar código de barra
                    if codigo_barra is None:
                        codigo_barra = ''
                    else:
                        codigo_barra = str(codigo_barra).strip()
                        if codigo_barra in ('None', 'nan', ''):
                            codigo_barra = ''
                    
                    # Validaciones
                    if cantidad < 0:
                        errores.append(f"Fila {row_idx}: La cantidad no puede ser negativa")
                        continue
                    
                    if costo < 0:
                        errores.append(f"Fila {row_idx}: El precio de compra no puede ser negativo")
                        continue
                    
                    if precio < 0:
                        errores.append(f"Fila {row_idx}: El precio unitario no puede ser negativo")
                        continue
                    
                    # Convertir a Decimal para Django
                    costo_decimal = Decimal(str(costo)).quantize(Decimal('0.01'))
                    precio_decimal = Decimal(str(precio)).quantize(Decimal('0.01'))
                    
                    # Buscar si ya existe un producto con ese código de barra
                    producto = None
                    if codigo_barra:
                        producto = Producto.objects.filter(codigo_barra=codigo_barra).first()
                    
                    if producto:
                        # SOLO ACTUALIZAR EL NOMBRE (NO modificar nada más)
                        nombre_anterior = producto.nombre_producto
                        producto.nombre_producto = nombre
                        producto.usuario_actualizacion = request.user.id if request.user.is_authenticated else None
                        # NO actualizar: cantidad, costo, precio, estado
                        producto.save()
                        
                        productos_actualizados.append({
                            'codigo_barra': codigo_barra,
                            'nombre_anterior': nombre_anterior,
                            'nombre_nuevo': nombre
                        })
                    else:
                        # Crear nuevo producto (con todos los datos)
                        producto = Producto(
                            nombre_producto=nombre,
                            cantidad_producto=cantidad,
                            costo_producto=costo_decimal,
                            precio_unitario=precio_decimal,
                            codigo_barra=codigo_barra,
                            estado=True,
                            usuario_creacion=request.user.id if request.user.is_authenticated else None,
                            usuario_actualizacion=request.user.id if request.user.is_authenticated else None
                        )
                        producto.save()
                        productos_creados.append(nombre)
                        
                except (ValueError, InvalidOperation) as e:
                    errores.append(f"Fila {row_idx}: Error en datos numéricos - {str(e)}")
                except Exception as e:
                    errores.append(f"Fila {row_idx}: Error inesperado - {str(e)}")
        
        # Construir mensaje de respuesta
        mensaje = []
        if productos_creados:
            mensaje.append(f"{len(productos_creados)} productos creados")
        if productos_actualizados:
            mensaje.append(f"{len(productos_actualizados)} productos actualizados (solo nombre)")
        
        if errores:
            mensaje.append(f"{len(errores)} errores encontrados")
        
        return Response({
            'message': ' - '.join(mensaje) if mensaje else 'No se procesaron productos',
            'creados': len(productos_creados),
            'actualizados': len(productos_actualizados),
            'errores': len(errores),
            'detalles': errores[:20],
            'lista_actualizados': productos_actualizados
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {'error': f'Error al leer el archivo: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )